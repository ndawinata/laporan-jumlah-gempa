from flask import Flask, render_template, request, make_response, session, redirect, url_for, flash, abort, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from DateTime import DateTime
import datetime
import os

app = Flask(__name__)
app.secret_key = 'randomterserah'

# upload gambar
app.config['UPLOAD_FOLDER'] = 'static'


@app.route('/', methods=['GET','POST'])
def uploadFile():
    if request.method == 'POST':
        
        if request.form['submit_button'] == 'Load': 
            name = request.form['file_name']
            
            dat = os.listdir('./static/file/data')
            
            return render_template('index.html', new_name=name, show="d-none", data=dat, path_download=name)
        
        if request.form['submit_button'] == 'New': 
            name = request.form['newfile'] + '.xlsx'
            path = "./static/file/data/" + name
            
            wb = load_workbook(filename="./static/file/templatexls/Laporan Jumlah Gempa.xlsx" )
            
            ws = wb.active
            
            wb.copy_worksheet(ws)
            wb.save(filename=path)
            
            dat = os.listdir('./static/file/data')
            
            return render_template('index.html', new_name=name, show="d-none", data=dat, path_download=name)
        
        if request.form['submit_button'] == 'Save':
            path = "./static/file/data/" + request.form['file_name']
            
            wb = load_workbook(filename=path )

            ws = wb.active
            
            x = request.form['textarea']
            
            if(ws.max_row == 2):
                y=0
            else:
                y=ws.max_row-2
            text = x.split(",")[0]
            waktu = (x.split(",")[1]).split(' ')
            mag = (text.split(":")[1]).split(' ')[0]
            tgl = DateTime(waktu[1] +' '+ waktu[2] + ' Asia/Jakarta').toZone('UTC').Date()
            wkt = DateTime(waktu[1] +' '+ waktu[2] + ' Asia/Jakarta').toZone('UTC').Time()
            lat = (x.split(", ")[2]).split(":")[1]
            lon = (x.split(",")[3]).split(' ')[0] 
            dep = (x.split(",")[4]).split(":")[1]
            ket = ((x.split(",")[3]).split('(')[1]).split(')')[0]
            ras = ''
            if(len(x.split(', '))==5):
                ras = (x.split(', ')[4]).split('::')[0]

            if lat.split(" ")[1] == 'LS' :
                lat = "-" + lat.split(" ")[0]
            else:
                lat = lat.split(" ")[0]
                
            y+=1
        
            data = [y, tgl, wkt, lat, lon, dep.split(" ")[0], mag, ket, ras]
            ws.append(data)
        
            wb.save(filename=path)
            
            dat = os.listdir('./static/file/data')
            
            return render_template('index.html', msg="Berhasil menyimpan data", show=" ", data=dat)
        
    dat = os.listdir('./static/file/data')
                
    return render_template('index.html', show="d-none", data=dat)

@app.errorhandler(404)
def not_found(error=None):
    message = {
        'status':404,
        'message': 'Not Found ' + request.url 
    }
    resp = jsonify(message)
    resp.status_code = 404
    return resp

if __name__ == "__main__":
    app.run(debug=True)