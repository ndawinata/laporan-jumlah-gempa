from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from DateTime import DateTime

wb = load_workbook(filename="Laporan Jumlah Gempa.xlsx")

# wb = Workbook()
ws = wb.active

if(ws.max_row == 2):
    y=0
else:
    y=ws.max_row-2

while True:
    x = input('masukkan: ')

    text = x.split(",")[0]
    waktu = (x.split(",")[1]).split(' ')
    mag = (text.split(":")[1]).split(' ')[0]
    tgl = DateTime(waktu[1] +' '+ waktu[2] + ' Asia/Jakarta').toZone('UTC').Date()
    wkt = DateTime(waktu[1] +' '+ waktu[2] + ' Asia/Jakarta').toZone('UTC').Time()
    lat = (x.split(", ")[2]).split(":")[1]
    lon = (x.split(",")[3]).split(' ')[0] 
    dep = (x.split(",")[4]).split(":")[1]
    ket = ((x.split(",")[3]).split('(')[1]).split(')')[0]
    ras = ''
    if(len(x.split(', '))==5):
        ras = (x.split(', ')[4]).split('::')[0]

    if lat.split(" ")[1] == 'LS' :
        lat = "-" + lat.split(" ")[0]
    else:
        lat = lat.split(" ")[0]
        
    y+=1
    
    data = [y, tgl, wkt, lat, lon, dep.split(" ")[0], mag, ket, ras]
    
    ws.append(data)
    
    wb.save(filename="Laporan Jumlah Gempa.xlsx")